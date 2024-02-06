from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

loc = './test.xlsx'

wb = load_workbook(loc)
ws = wb.active

maxRows = ws.max_row
maxCols = ws.max_column

colHeads = ("Name: ", "Type: ", "Location: ", "Mobile: ")

for row in range(maxRows + 1, maxRows + 2):
    for col in range(1, maxCols + 1):
        letter = get_column_letter(col)
        ws[letter + str(row)].value = input(colHeads[col - 1])

wb.save('./test.xlsx')
